from __future__ import annotations

import csv
import io
import inspect
import queue
import sched
import threading
import time
import os
import sys
from collections import defaultdict
from datetime import datetime
from dataclasses import dataclass
from multiprocessing import Process
from pathlib import Path
from typing import TYPE_CHECKING, Any, Dict, Iterable, List

from console_utils import ensure_own_console
from console_utils import ensure_own_console
from config import (
    SCHEDULE_TABLE_PATH,
    SCHEDULE_CONCURRENCY,
    CHROME_USER_DATA_DIR,
    SCHEDULE_SHOW_CONSOLE,
)

if TYPE_CHECKING:
    from social_poster import MediumJobConfig, RunnerConfig

CSV_PATH = Path(SCHEDULE_TABLE_PATH)
ENCODING_CANDIDATES: tuple[str, ...] = (
    "utf-8-sig",
    "utf-8",
    "utf-16",
    "cp1258",
    "latin-1",
)
DEFAULT_LIMIT = SCHEDULE_CONCURRENCY
DEFAULT_SHOW_CONSOLE = bool(SCHEDULE_SHOW_CONSOLE)
SCHEDULE_TIME_FORMATS: tuple[str, ...] = (
    "%H:%M",
    "%H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d %H:%M:%S",
)
try:  # optional Excel support
    import pandas as pd  # type: ignore
except Exception:  # pragma: no cover
    pd = None
try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None


def _log(message: str) -> None:
    caller = inspect.currentframe().f_back  # type: ignore[assignment]
    line = caller.f_lineno if caller else -1
    pid = os.getpid()
    formatted = f"[pid {pid:>6}] [line {line:04d}] {message}"
    encoding = getattr(sys.stdout, "encoding", None) or "utf-8"
    try:
        sys.stdout.buffer.write((formatted + "\n").encode(encoding, errors="replace"))
        sys.stdout.flush()
    except Exception:
        print(formatted)


def _preview(value: str, limit: int = 30) -> str:
    value = value or ""
    if len(value) <= limit:
        return value
    return value[: limit - 3] + "..."


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return text.strip()


_MOJIBAKE_MARKERS = ("Ã", "Â", "Ð", "‰", "Ê", "¤", "�")


def _normalize_field(value: Any) -> str:
    text = _clean(value)
    if not text:
        return text
    if any(marker in text for marker in _MOJIBAKE_MARKERS):
        try:
            repaired = text.encode("latin-1", errors="strict").decode("utf-8")
            return repaired.strip()
        except Exception:
            return text
    return text

_MOJIBAKE_MARKERS = ("Ã", "Â", "Ð", "Ê", "¤", "�")


def _parse_schedule_timestamp(value: str, date_hint: str | None = None) -> datetime | None:
    text = (value or "").strip()
    if not text:
        return None
    now = datetime.now()
    date_obj: datetime | None = None
    if date_hint:
        for fmt in ("%d/%m", "%d-%m", "%d.%m"):
            try:
                parsed_date = datetime.strptime(date_hint.strip(), fmt)
                date_obj = parsed_date.replace(year=now.year)
                break
            except ValueError:
                continue
    for fmt in SCHEDULE_TIME_FORMATS:
        try:
            parsed = datetime.strptime(text, fmt)
        except ValueError:
            continue
        if date_obj:
            parsed = parsed.replace(
                year=date_obj.year,
                month=date_obj.month,
                day=date_obj.day,
            )
        elif "%Y" not in fmt and "%y" not in fmt and "%m" not in fmt and "%d" not in fmt:
            parsed = parsed.replace(year=now.year, month=now.month, day=now.day)
        return parsed
    return None


def _is_xlsx(path: Path) -> bool:
    lower = path.name.lower()
    if lower.endswith((".xlsx", ".xls")):
        return True
    try:
        with path.open("rb") as fh:
            return fh.read(2) == b"PK"
    except Exception:
        return False


def _read_csv_records(path: Path) -> List[Dict[str, Any]]:
    data = path.read_bytes()
    text: str | None = None
    for encoding in ENCODING_CANDIDATES:
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = data.decode("latin-1", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def _read_excel_records(path: Path) -> List[Dict[str, Any]]:
    if pd is not None:
        df = pd.read_excel(path)  # type: ignore[arg-type]
        return df.fillna("").to_dict(orient="records")
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for Excel schedules")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    records: List[Dict[str, Any]] = []
    for row in rows[1:]:
        record: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[idx] if idx < len(row) else ""
        records.append(record)
    return records


def read_schedule(path: Path = CSV_PATH) -> Dict[str, List[str]]:
    path = path.expanduser()
    if path.suffix.lower() in (".xlsx", ".xls") or _is_xlsx(path):
        raw_rows = _read_excel_records(path)
        _log(f"INFO:READ_SCHEDULE excel rows={len(raw_rows)} file={path}")
    else:
        raw_rows = _read_csv_records(path)
        _log(f"INFO:READ_SCHEDULE csv rows={len(raw_rows)} file={path}")

    fields: Dict[str, List[str]] = {
        "platform": [],
        "profile": [],
        "type": [],
        "title": [],
        "content": [],
        "images": [],
        "schedule_time": [],
        "schedule_date": [],
        "link": [],
        "__row_index": [],
    }

    for idx, row in enumerate(raw_rows, start=1):
        for key in fields:
            if key == "profile":
                value = (
                    row.get("profile")
                    or row.get("profile_path")
                    or row.get("email")
                    or ""
                )
            elif key == "schedule_date":
                value = row.get("schedule_date") or row.get("date") or ""
            elif key == "__row_index":
                value = str(idx + 1)  # include header row
            else:
                value = row.get(key, "")
            fields[key].append(_normalize_field(value))

    times_preview = ", ".join(_preview(t, 8) or "-" for t in fields["schedule_time"])
    dates_preview = ", ".join(_preview(d, 5) or "-" for d in fields["schedule_date"])
    _log(f"INFO:SCHEDULE_TIMES {times_preview or '<none>'} | DATES {dates_preview or '<none>'}")
    return fields


def build_jobs(columns: Dict[str, List[str]]) -> List["ScheduleJob"]:
    _log("INFO:BUILD_JOBS_START")
    count = max((len(values) for values in columns.values()), default=0)
    jobs: List[ScheduleJob] = []
    for idx in range(count):
        data = {key: columns[key][idx] if idx < len(columns[key]) else "" for key in columns}
        if not any(data.values()):
            continue
        if _normalize_field(data.get("link", "")):
            _log(
                f"INFO:SKIP_JOB row={data.get('__row_index','?')} profile={data.get('profile','')} reason=link_present"
            )
            continue
        jobs.append(ScheduleJob.from_dict(data))
    _log(f"INFO:BUILD_JOBS total={len(jobs)}")
    for idx, job in enumerate(jobs, start=1):
        _log(
            f"INFO:JOB_SUMMARY #{idx} platform={job.platform} date={(job.schedule_date or 'today')!r} time={(job.schedule_time or 'imm')!r} title='{_preview(job.title)}' content='{_preview(job.content)}'"
        )
    return jobs


@dataclass
class ScheduleJob:
    platform: str
    profile: str
    type: str
    title: str
    content: str
    images: str
    schedule_time: str
    schedule_date: str
    link: str
    row_index: int
    table_path: Path | None = None

    @classmethod
    def from_dict(cls, row: Dict[str, str]) -> "ScheduleJob":
        return cls(
            platform=_normalize_field(row.get("platform", "")) or "Medium",
            profile=_normalize_field(row.get("profile", "")),
            type=_normalize_field(row.get("type", "")) or "medium",
            title=_normalize_field(row.get("title", "")) or "Untitled",
            content=_normalize_field(row.get("content", "")),
            images=_normalize_field(row.get("images", "")),
            schedule_time=_normalize_field(row.get("schedule_time", "")),
            schedule_date=_normalize_field(row.get("schedule_date", "")),
            link=_normalize_field(row.get("link", "")),
            row_index=int(_normalize_field(row.get("__row_index", "0")) or "0"),
        )

    def to_runner_config(self) -> "RunnerConfig":
        from social_poster import MediumJobConfig, RunnerConfig

        platform = self.platform or "Medium"
        if platform.lower() != "medium":
            return RunnerConfig(platform=platform)
        profile_name = self.profile or "Default"
        medium_cfg = MediumJobConfig(
            profile_path=self.resolve_profile_path(),
            profile_name=profile_name,
            title=self.title or "Untitled",
            content=self.content or "",
            schedule_table=str(self.table_path) if self.table_path else None,
            schedule_row=self.row_index,
        )
        return RunnerConfig(platform="Medium", medium=medium_cfg)

    def resolve_profile_path(self) -> str:
        base = Path(CHROME_USER_DATA_DIR).expanduser()
        profiles_root = Path(r"D:\TOOL\social-poster\profiles")
        value = (self.profile or "").strip()
        if not value:
            return str(base)
        path = Path(value).expanduser()
        if not path.is_absolute():
            path = profiles_root / value
        if not str(path).lower().startswith(str(profiles_root).lower()):
            path = profiles_root / path.name
        return str(path)


def _run_single_job(job: ScheduleJob, show_console: bool = False) -> None:
    from social_poster import run_job_inline

    cfg = job.to_runner_config()
    console_label = job.profile or job.platform or "job"
    events = run_job_inline(
        cfg,
        open_console=show_console,
        console_title=f"{console_label.strip() or 'default'}",
    )

    _log(
        f"INFO:LAUNCH_JOB platform={job.platform} time='{job.schedule_time}' title='{_preview(job.title)}'"
    )
    publish_url: str | None = None
    for level, message in events:
        _log(f"LOG:{level.upper()} {message}")
        if level == "success" and "Medium URL:" in message:
            publish_url = message.split("Medium URL:", 1)[-1].strip()
    if publish_url and job.table_path and job.row_index:
        try:
            write_link_to_schedule(Path(job.table_path), job.row_index, publish_url)
            _log(
                f"INFO:LINK_UPDATE row={job.row_index} url='{publish_url}' table='{job.table_path}'"
            )
        except Exception as exc:
            _log(
                f"WARN:LINK_UPDATE_FAILED row={job.row_index} table='{job.table_path}' err={exc}"
            )



def _group_jobs_by_profile(jobs: List[ScheduleJob]) -> Dict[str, List[ScheduleJob]]:
    grouped: Dict[str, List[ScheduleJob]] = defaultdict(list)
    for job in jobs:
        resolved = job.resolve_profile_path()
        key = resolved.strip().lower() or "default"
        grouped[key].append(job)
    return grouped


def _profile_worker(group_id: str, jobs: List[ScheduleJob], show_console: bool) -> None:
    _ensure_process_console(group_id, show_console)
    for job in jobs:
        _log(
            "INFO:RUN_JOB "
            + f"profile={job.profile or 'N/A'} "
            + f"type={job.type or 'N/A'} "
            + f"title='{_preview(job.title)}' "
            + f"content='{_preview(job.content)}' "
            + f"images='{_preview(job.images)}' "
            + f"time='{job.schedule_time or 'imm'}' "
            + f"row={job.row_index} "
            + f"link='{job.link or ''}'"
        )
        _run_single_job(job, show_console=show_console)
    _log(f"INFO:PROFILE_WORKER finished profile={group_id}")

def _dispatch_time_slot(
    slot_label: str, jobs: List[ScheduleJob], limit: int, show_console: bool
) -> None:
    _log(f"INFO:TIME_SLOT_DISPATCH label={slot_label} jobs={len(jobs)}")
    grouped = _group_jobs_by_profile(jobs)
    processes: List[Process] = []
    for group_id, group_jobs in grouped.items():
        while True:
            alive = [p for p in processes if p.is_alive()]
            if len(alive) < limit:
                processes = alive
                break
            _log(f"INFO:PROFILE_MANAGER waiting for slot alive={len(alive)}/{limit}")
            time.sleep(5)
            processes = alive
        proc = Process(target=_profile_worker, args=(group_id, group_jobs, show_console))
        proc.start()
        _log(f"INFO:PROFILE_PROCESS start profile={group_id} pid={proc.pid} jobs={len(group_jobs)}")
        processes.append(proc)

    for proc in processes:
        proc.join()
        _log(f"INFO:PROFILE_PROCESS finished pid={proc.pid}")


def _group_jobs_by_time(jobs: List[ScheduleJob]) -> tuple[List[ScheduleJob], Dict[datetime, List[ScheduleJob]]]:
    immediate: List[ScheduleJob] = []
    scheduled: Dict[datetime, List[ScheduleJob]] = defaultdict(list)
    now = datetime.now()
    for job in jobs:
        target = _parse_schedule_timestamp(job.schedule_time, job.schedule_date)
        if target is None or target <= now:
            immediate.append(job)
        else:
            delay = max(0.0, (target - now).total_seconds())
            _log(
                f"INFO:JOB_SCHEDULE profile={job.profile or 'default'} title='{_preview(job.title)}' target={target:%Y-%m-%d %H:%M:%S} delay={int(delay)}s"
            )
            scheduled[target].append(job)
    if immediate:
        _log(
            "INFO:DISPATCH_IMMEDIATE "
            + ", ".join(
                f"{job.profile or 'default'}@{job.schedule_date or '--'}/{job.schedule_time or 'imm'}"
                for job in immediate
            )
        )
    for ts, slot in scheduled.items():
        label = ts.strftime("%Y-%m-%d %H:%M:%S")
        entries: list[str] = []
        for job in slot:
            delay = max(0.0, (ts - now).total_seconds())
            entries.append(
                f"{job.profile or 'default'}@{job.schedule_date or '--'}/{job.schedule_time or 'imm'} delay={int(delay)}s"
            )
        _log(f"INFO:TIME_SLOT_SCHEDULE label={label} jobs={len(slot)} list=[{'; '.join(entries)}]")
    return immediate, scheduled


def _ensure_process_console(group_id: str, enabled: bool) -> None:
    if not enabled:
        return
    try:
        ensure_own_console(f"Profile-{group_id}", verbose=True)
    except Exception as exc:  # pragma: no cover - best effort
        _log(f"WARN:CONSOLE_FAIL profile={group_id} err={exc}")


def write_link_to_schedule(table: Path, row_index: int, url: str) -> None:
    table = table.expanduser()
    if row_index <= 1:
        return
    if table.suffix.lower() in (".xlsx", ".xls"):
        if load_workbook is None:
            raise RuntimeError("openpyxl is required to update Excel schedules")
        wb = load_workbook(table)
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        link_col = None
        for idx, header in enumerate(headers, start=1):
            if header.lower() == "link":
                link_col = idx
                break
        if link_col is None:
            link_col = len(headers) + 1
            ws.cell(row=1, column=link_col, value="link")
        ws.cell(row=row_index, column=link_col, value=url)
        wb.save(table)
        return

    if not table.exists():
        raise FileNotFoundError(table)
    with table.open("r", newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.reader(fh))
    if not rows:
        return
    header = rows[0]
    header_lower = [h.strip().lower() for h in header]
    if "link" in header_lower:
        link_idx = header_lower.index("link")
    else:
        header.append("link")
        link_idx = len(header) - 1
        for row in rows[1:]:
            row.extend([""] * (len(header) - len(row)))
    while len(rows) < row_index:
        rows.append([""] * len(header))
    row = rows[row_index - 1]
    if len(row) < len(header):
        row.extend([""] * (len(header) - len(row)))
    row[link_idx] = url
    with table.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerows(rows)


def main(
    table: Path | None = None,
    limit: int | None = None,
    show_console: bool | None = None,
) -> None:
    table = (table or CSV_PATH).expanduser()
    limit = max(1, limit or DEFAULT_LIMIT)
    show_console = DEFAULT_SHOW_CONSOLE if show_console is None else show_console
    columns = read_schedule(table)
    jobs = build_jobs(columns)
    for job in jobs:
        job.table_path = table
    if not jobs:
        _log("WARN: No jobs found in schedule.")
        return
    immediate_jobs, scheduled_jobs = _group_jobs_by_time(jobs)
    # input("stop a second")
    if immediate_jobs:
        _dispatch_time_slot("immediate", immediate_jobs, limit, show_console)

    scheduler = sched.scheduler(time.time, time.sleep)
    for target, slot_jobs in sorted(scheduled_jobs.items(), key=lambda item: item[0]):
        delay = max(0.0, (target - datetime.now()).total_seconds())
        label = target.strftime("%Y-%m-%d %H:%M:%S")
        _log(
            f"INFO:TIME_SLOT_REGISTER label={label} jobs={len(slot_jobs)} delay={int(delay)}s show_console={show_console}"
        )
        scheduler.enter(
            delay,
            1,
            _dispatch_time_slot,
            argument=(label, slot_jobs, limit, show_console),
        )

    if scheduled_jobs:
        scheduler.run()


if __name__ == "__main__":
    main()
